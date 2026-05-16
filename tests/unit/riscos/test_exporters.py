from io import BytesIO

import pytest
from openpyxl import load_workbook

from src.riscos.exporters import (
    _nome_arquivo,
    _primeiro_plano_acao,
    _texto,
    exportar_risco_excel,
    exportar_risco_pdf,
    exportar_riscos_excel,
)


@pytest.mark.django_db
class TestHelpersExportacao:
    def test_texto_substitui_none_por_traco(self):
        # este helper deve evitar valores nulos nas exportacoes
        assert _texto(None) == "-"
        assert _texto("valor") == "valor"

    def test_nome_arquivo_monta_sufixo_opcional(self):
        # o nome do arquivo deve incluir identificador apenas quando necessario
        assert _nome_arquivo("planos-risco", "xlsx") == "planos-risco.xlsx"
        assert _nome_arquivo("plano-risco", "pdf", 10) == "plano-risco-10.pdf"

    def test_primeiro_plano_acao_retorna_none_sem_plano(self, risco_basico):
        # quando nao existe tratamento o helper deve retornar none
        assert _primeiro_plano_acao(risco_basico) is None


@pytest.mark.django_db
class TestExportadores:
    def test_exporta_lista_excel_com_colunas_principais(self, risco_com_plano):
        # a exportacao da listagem deve gerar um workbook valido com cabecalho conhecido
        response = exportar_riscos_excel([risco_com_plano])

        assert response.status_code == 200
        assert response["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert 'filename="planos-risco.xlsx"' in response["Content-Disposition"]

        workbook = load_workbook(filename=BytesIO(response.content))
        planilha = workbook.active

        assert planilha["A1"].value == "ID"
        assert planilha["B1"].value == "Unidade/Departamento"
        assert planilha["Q2"].value == risco_com_plano.nivel_residual
        assert planilha["R2"].value == "Mitigar"

    def test_exporta_plano_individual_excel_sem_plano_acao(self, risco_basico):
        # o excel individual precisa preencher tracos quando nao ha tratamento
        response = exportar_risco_excel(risco_basico)

        assert response.status_code == 200
        assert f'filename="plano-risco-{risco_basico.id}.xlsx"' in response["Content-Disposition"]

        workbook = load_workbook(filename=BytesIO(response.content))
        planilha = workbook.active
        linhas = {planilha[f"A{i}"].value: planilha[f"B{i}"].value for i in range(2, planilha.max_row + 1)}

        assert linhas["Tipo de resposta"] == "-"
        assert linhas["Responsavel"] == "-"
        assert linhas["Status da acao"] == "-"

    def test_exporta_plano_individual_pdf(self, risco_com_plano):
        # o pdf individual deve retornar um arquivo pdf valido
        response = exportar_risco_pdf(risco_com_plano)

        assert response.status_code == 200
        assert response["Content-Type"] == "application/pdf"
        assert f'filename="plano-risco-{risco_com_plano.id}.pdf"' in response["Content-Disposition"]
        assert response.content.startswith(b"%PDF")
