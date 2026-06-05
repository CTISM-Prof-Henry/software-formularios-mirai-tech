from datetime import date
from io import BytesIO
from xml.sax.saxutils import escape

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _texto(valor):
    if valor is None:
        return "-"
    return str(valor)


def _paragrafo(valor, estilo):
    return Paragraph(escape(_texto(valor)), estilo)


def _primeiro_plano_acao(risco):
    return next(iter(risco.planos_acao.all()), None)


def _nome_arquivo(prefixo, extensao, identificador=None):
    sufixo = f"-{identificador}" if identificador else ""
    return f"{prefixo}{sufixo}.{extensao}"


def _resposta_arquivo(conteudo, content_type, filename):
    response = HttpResponse(conteudo, content_type=content_type)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _configurar_planilha(ws):
    header_fill = PatternFill("solid", fgColor="0A376D")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(_texto(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 48)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def exportar_riscos_excel(riscos):
    wb = Workbook()
    ws = wb.active
    ws.title = "Planos de Risco"

    headers = [
        "ID",
        "Unidade/Departamento",
        "Categoria",
        "Desafio PDI",
        "Objetivo PDI",
        "Macroprocesso",
        "Evento",
        "Causa",
        "Consequencia",
        "Controles atuais",
        "Eficacia dos controles",
        "Probabilidade",
        "Impacto",
        "Nivel inerente",
        "Probabilidade residual",
        "Impacto residual",
        "Nivel residual",
        "Tipo de resposta",
        "Responsavel",
        "Status da acao",
        "Descricao da acao",
    ]
    ws.append(headers)

    for risco in riscos:
        plano_acao = _primeiro_plano_acao(risco)
        desafio = risco.objetivo.desafio
        ws.append([
            risco.id,
            risco.setor.label_curto,
            risco.categoria,
            f"Desafio {desafio.numero} - {desafio.descricao}",
            f"{risco.objetivo.codigo} - {risco.objetivo.descricao}",
            risco.macroprocesso.nome,
            risco.evento,
            risco.causa,
            risco.consequencia,
            risco.controles_atuais,
            risco.eficacia_controle,
            risco.probabilidade,
            risco.impacto,
            risco.nivel_risco,
            risco.prob_residual,
            risco.imp_residual,
            risco.nivel_residual,
            plano_acao.tipo_resposta if plano_acao else "-",
            plano_acao.responsavel if plano_acao else "-",
            plano_acao.status if plano_acao else "-",
            plano_acao.descricao_acao if plano_acao else "-",
        ])

    _configurar_planilha(ws)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return _resposta_arquivo(
        buffer.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        _nome_arquivo("planos-risco", "xlsx"),
    )


def exportar_risco_excel(risco):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Plano {risco.id}"

    plano_acao = _primeiro_plano_acao(risco)
    desafio = risco.objetivo.desafio
    dados = [
        ("ID", risco.id),
        ("Unidade/Departamento", risco.setor.label_curto),
        ("Categoria", risco.categoria),
        ("Desafio PDI", f"Desafio {desafio.numero} - {desafio.descricao}"),
        ("Objetivo PDI", f"{risco.objetivo.codigo} - {risco.objetivo.descricao}"),
        ("Macroprocesso", risco.macroprocesso.nome),
        ("Evento", risco.evento),
        ("Causa", risco.causa),
        ("Consequencia", risco.consequencia),
        ("Controles atuais", risco.controles_atuais),
        ("Eficacia dos controles", risco.eficacia_controle),
        ("Nivel inerente", f"{risco.nivel_risco} (P{risco.probabilidade} x I{risco.impacto})"),
        ("Nivel residual", f"{risco.nivel_residual} (P{risco.prob_residual} x I{risco.imp_residual})"),
        ("Tipo de resposta", plano_acao.tipo_resposta if plano_acao else "-"),
        ("Responsavel", plano_acao.responsavel if plano_acao else "-"),
        ("Status da acao", plano_acao.status if plano_acao else "-"),
        ("Descricao da acao", plano_acao.descricao_acao if plano_acao else "-"),
        ("Parceiros", plano_acao.parceiros if plano_acao else "-"),
        ("Periodo", f"{plano_acao.data_inicio} ate {plano_acao.data_fim}" if plano_acao else "-"),
        ("Observacoes", plano_acao.observacoes if plano_acao else "-"),
    ]

    ws.append(["Campo", "Valor"])
    for linha in dados:
        ws.append(linha)

    _configurar_planilha(ws)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return _resposta_arquivo(
        buffer.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        _nome_arquivo("plano-risco", "xlsx", risco.uuid),
    )


def exportar_relatorio_gerencial(riscos):
    riscos = list(riscos)
    hoje = date.today()

    nivel_counts = {"Extremo": 0, "Alto": 0, "Moderado": 0, "Baixo": 0}
    categoria_counts = {}
    proximas_acoes = []

    for risco in riscos:
        n = risco.nivel_residual
        if n >= 20:
            nivel_counts["Extremo"] += 1
        elif n >= 12:
            nivel_counts["Alto"] += 1
        elif n >= 4:
            nivel_counts["Moderado"] += 1
        else:
            nivel_counts["Baixo"] += 1

        categoria_counts[risco.categoria] = categoria_counts.get(risco.categoria, 0) + 1

        plano = next(iter(risco.planos_acao.all()), None)
        if plano and plano.status != "Concluída":
            dias = (plano.data_fim - hoje).days
            if dias <= 14:
                proximas_acoes.append((dias, risco, plano))

    proximas_acoes.sort(key=lambda x: x[0])

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Relatório Gerencial de Riscos", styles["Title"]))
    story.append(Paragraph(f"Gerado em {hoje.strftime('%d/%m/%Y')}", styles["Normal"]))
    story.append(Spacer(1, 20))

    story.append(Paragraph("<b>Resumo Geral</b>", styles["Heading2"]))
    resumo_dados = [
        ["Total de planos", str(len(riscos))],
        ["Riscos Extremos", str(nivel_counts["Extremo"])],
        ["Riscos Altos", str(nivel_counts["Alto"])],
        ["Riscos Moderados", str(nivel_counts["Moderado"])],
        ["Riscos Baixos", str(nivel_counts["Baixo"])],
    ]
    t = Table(resumo_dados, colWidths=[200, 290])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EEF4FB")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#0A376D")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9E2EC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 20))

    story.append(Paragraph("<b>Distribuição por Categoria</b>", styles["Heading2"]))
    cat_dados = [["Categoria", "Quantidade"]] + [[k, str(v)] for k, v in sorted(categoria_counts.items())]
    tc = Table(cat_dados, colWidths=[200, 290])
    tc.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A376D")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9E2EC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(tc)
    story.append(Spacer(1, 20))

    if proximas_acoes:
        story.append(Paragraph("<b>Ações com prazo próximo ou vencido (próximos 14 dias)</b>", styles["Heading2"]))
        acao_dados = [["Unidade", "Evento", "Responsável", "Prazo", "Status"]]
        for dias, risco, plano in proximas_acoes[:10]:
            prazo_str = plano.data_fim.strftime("%d/%m/%Y")
            if dias < 0:
                prazo_str += " (vencido)"
            acao_dados.append([
                risco.setor.label_curto[:30],
                risco.evento[:40],
                plano.responsavel[:25],
                prazo_str,
                plano.status,
            ])
        ta = Table(acao_dados, colWidths=[100, 160, 90, 70, 70])
        ta.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A376D")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9E2EC")),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("WORDWRAP", (0, 0), (-1, -1), True),
        ]))
        story.append(ta)

    doc.build(story)
    buffer.seek(0)
    return _resposta_arquivo(
        buffer.getvalue(),
        "application/pdf",
        _nome_arquivo("relatorio-gerencial", "pdf"),
    )


def exportar_risco_pdf(risco):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"Plano de Risco #{risco.id}", styles["Title"]),
        Paragraph(risco.setor.label_curto, styles["Normal"]),
        Spacer(1, 16),
    ]

    plano_acao = _primeiro_plano_acao(risco)
    desafio = risco.objetivo.desafio
    dados = [
        ("Categoria", risco.categoria),
        ("Desafio PDI", f"Desafio {desafio.numero} - {desafio.descricao}"),
        ("Objetivo PDI", f"{risco.objetivo.codigo} - {risco.objetivo.descricao}"),
        ("Macroprocesso", risco.macroprocesso.nome),
        ("Evento de risco", risco.evento),
        ("Causa", risco.causa),
        ("Consequencia", risco.consequencia),
        ("Controles atuais", risco.controles_atuais),
        ("Eficacia dos controles", risco.eficacia_controle),
        ("Nivel inerente", f"{risco.nivel_risco} (P{risco.probabilidade} x I{risco.impacto})"),
        ("Nivel residual", f"{risco.nivel_residual} (P{risco.prob_residual} x I{risco.imp_residual})"),
        ("Tipo de resposta", plano_acao.tipo_resposta if plano_acao else "-"),
        ("Responsavel", plano_acao.responsavel if plano_acao else "-"),
        ("Status da acao", plano_acao.status if plano_acao else "-"),
        ("Descricao da acao", plano_acao.descricao_acao if plano_acao else "-"),
        ("Periodo", f"{plano_acao.data_inicio} ate {plano_acao.data_fim}" if plano_acao else "-"),
    ]

    tabela = Table(
        [[Paragraph(f"<b>{campo}</b>", styles["BodyText"]), _paragrafo(valor, styles["BodyText"])] for campo, valor in dados],
        colWidths=[130, 360],
    )
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EEF4FB")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#0A376D")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9E2EC")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(tabela)

    doc.build(story)
    buffer.seek(0)
    return _resposta_arquivo(
        buffer.getvalue(),
        "application/pdf",
        _nome_arquivo("plano-risco", "pdf", risco.uuid),
    )
